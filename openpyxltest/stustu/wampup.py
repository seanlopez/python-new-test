from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("testplatform.xlsx")
sheet = wb.sheetnames
print(sheet)
sheetname = sheet[0]
ws = wb[sheetname]
ws.append([])
ws.append(["tian", "qi", "Sean"])
wb.save("testplatform.xlsx")
