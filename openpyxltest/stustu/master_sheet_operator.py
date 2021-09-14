from openpyxl import Workbook
from openpyxl import load_workbook
import master_sheet_format


class MasterSheet(object):
    def __init__(self, file_path, sheet_name, override=False):
        """
        create or load a master_sheet.

        :param file_path: the file path of the master_sheet
        :param sheet_name: the work sheet name of the excel work book
        :param override: this parameter could be "False" or "True".
            for "True", if the specified file is existed, it will be deleted and create a new one.
            for "False", the master sheet (as an excel worksheet) will be inserted in the current file.
        """
        self.sheet_name = sheet_name
        self.file_path = file_path
        if override == False:
            self.wb = load_workbook(file_path+sheet_name)  #Load the existing workbook
        else:
            self.wb = Workbook()    #create a new workbook in RAM

        self.fm = master_sheet_format.master_sheet_format()

    def add_section(self, section_name, header0, header1, override=True):
        """
        add a new section into the master sheet

        :param section_name: the section name.   each section should be identified by an open flag an a close flag,
            and the section name should be located in the open flag which is a excel cell with a particular color.
        :param header0: the column names for the first table which is used to save date for original devices.
            this parameter should be a list consisted with column names.
        :param header1: the column names for the second table which is used to save date for new devices.
            this parameter should be a list consisted with column names.
        :param override: this parameter could be "False" or "True".
            for "True", if the section is existed, remove it by invoking <remove_section>, then re-create the section.
            for "False", terminate this task and raise a particular exception or just return None.
        :return: return a Non-zero value if this task was finished successfully, otherwise, return None.
        """
        sheet_names = self.wb.sheetnames
        num1_sheet = sheet_names[0]
        ws = self.wb[num1_sheet]
        colA = ws["A"]
        worksheet_start_row_num = str(len(colA)+2)
        section_tag_cell_num = "A" + worksheet_start_row_num
        value_col_A = []
        for cell in colA: value_col_A.append(cell.value)
        if section_name in value_col_A:
            if override == False:
                print("Warning! It's an existing section in worksheet")
                return None
            else:
                pass
        else:
            for ws_name in sheet_names:
                ws = self.wb[ws_name]
                ws[section_tag_cell_num] = section_name
                section_tag_cell = ws[section_tag_cell_num]
                ws.append(header0)
                max_row1 = ws[str(ws.max_row)]
                print(max_row1)
                ws.append([])
                ws.append(header1)
                max_row2 = ws[str(ws.max_row)]
                print(max_row2)
                last_cell_num = "A" + str(ws.max_row+1)
                ws[last_cell_num] = "End"
                section_end_tag = ws[last_cell_num]
                self.fm.section_format(section_tag_cell, section_end_tag)
                self.fm.section_title_format(max_row1, max_row2)
                self.fm.oversheet_format(ws)

    def section_format(self, section_tag_cell, section_end_tag):
        pass


    def remove_section(self, section_name):
        """
        remove the specified section from the master sheet.

        :param section_name: the section name
        :return: return a Non-zero value if this task was finished successfully, otherwise, return None.
        """
        pass

    def erase_table(self, section_name, table_index):
        """
        erase all the data records of the specified table of the particular section.
            Note: just erase the data of the table but not the table header

        :param section_name: the section name
        :param table_index: the table index. (each section only contain 2 tables with index 0, 1)
        :return: return the number of the rows that were erased. if it's failed, return a negative integer.
        """
        pass

    @property
    def section_names(self):
        """
        iterate through the master sheet and collect the section names in this master sheet.

        :return: a list that contains each section names. if there is no section, return a empty list.
        """
        pass

    def get_section(self, section_name):
        """
        get the data of the specified section and return the data as a list

        :param section_name: section name
        :return: a nested list which representing the tables.  the following is an example:
            [
                # the first table in the section
                [
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                        ... ...
                    {"column1": value1, "column2": value2, ... "columnN": valueN}
                ],
                # the second table in the section
                [
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                        ... ...
                    {"column1": value1, "column2": value2, ... "columnN": valueN}
                ]
            ]
        """
        pass

    @property
    def master_sheet(self):
        """
        get the whole data of the master sheet.
        Note: (for gathering this data, we can iterate the section_names by
            method <get_section>

        :return: a dictionary contains each section data. the following is an example:
            {
                "section1_name": section1_data,
                "section2_name": section2_data,
                    ... ...
                "sectionN_name": sectionN_data

            }

            Note: for the structure of section_data, please refer to the return value of method <get_section>
        """
        pass

    def get_table(self, section_name, table_index):
        """
        get the data of the specified table.
        Note: (for gathering this data, we can extract data from the return value of
            method <get_section>)

        :param section_name: the section name which the table is located in.
        :param table_index: the table index we want to get.
        :return: a list of dictionaries which presenting the data records, the following is an example:
             [
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                    ... ...
                {"column1": value1, "column2": value2, ... "columnN": valueN}
            ]
        """
        pass

    def insert(self, section_name, table_index, data):
        """
        insert data into a particular table of the specified section.

        :param section_name: section name.
        :param table_index: table index. (each section only contain 2 tables with index 0, 1)
        :param data: the data need to be inserted. it should be a dictionary that presents the data record (a row)
            the structure of the data should like the following example:
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                    ... ...
                {"column1": value1, "column2": value2, ... "columnN": valueN}
        :return: the number of the rows that were inserted. if failed, raise a exception or return a negative integer.
        """
        pass

    def bulk_insert(self, section_name, table_index, data_list, mode="merge"):
        """
        insert bulk data into the specified table.

        :param section_name: the section name.
        :param table_index: the table index. (each section only contain 2 tables with index 0, 1)
        :param data_list: a list of dictionaries. the following is an example:
            [
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                    ... ...
                {"column1": value1, "column2": value2, ... "columnN": valueN}
            ]
        :param mode: this parameter could be "merge" or "override".
            for merge, add the data into the existing table.
            for override, erase the existing table first, then insert the data. .
        :return: the number of the rows that were inserted. if failed, raise a exception or return a negative integer.
        """
        pass
