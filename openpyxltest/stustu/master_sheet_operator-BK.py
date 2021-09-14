import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import master_sheet_format


class MasterSheet(object):
    def __init__(self, file_path, sheet_name, override=False):
        """
        create or load a master_sheet.

        :param file_path: the file path of the master_sheet
        :param sheet_name: the work sheet name of the excel work book
        :param override: this parameter could be "False" or "True".
            for "True", if the specified file is existed, it will be deleted and create a new one.
            for "False", the master sheet (as an excel worksheet) will be inserted in the current file.
        """
        self.sheet_name = sheet_name
        self.file_path = file_path
        if override == False:
            self.wb = load_workbook(file_path+sheet_name)  #Load the existing workbook
        else:
            self.wb = Workbook()    #create a new workbook in RAM

        self.sheet_names = self.wb.sheetnames    #load all sheets in workbook
        self.fm = master_sheet_format.master_sheet_format()    #sheet format editor

    def add_section(self, section_name, header0, header1, override=True):
        """
        add a new section into the master sheet

        :param section_name: the section name.   each section should be identified by an open flag an a close flag,
            and the section name should be located in the open flag which is a excel cell with a particular color.
        :param header0: the column names for the first table which is used to save date for original devices.
            this parameter should be a list consisted with column names.
        :param header1: the column names for the second table which is used to save date for new devices.
            this parameter should be a list consisted with column names.
        :param override: this parameter could be "False" or "True".
            for "True", if the section is existed, remove it by invoking <remove_section>, then re-create the section.
            for "False", terminate this task and raise a particular exception or just return None.
        :return: return a Non-zero value if this task was finished successfully, otherwise, return None.
        """
        num1_sheet = self.sheet_names[0]
        ws = self.wb[num1_sheet]  # load the first sheet in workbook
        colA = ws["A"]
        worksheet_start_row_num = str(len(colA) + 2)  # add tag row number
        section_tag_cell_num = "A" + worksheet_start_row_num  # section tag cell position
        value_col_A = []
        for cell in colA: value_col_A.append(cell.value)
        if section_name in value_col_A:
            if override == False:
                print("Warning! It's an existing section in worksheet")
                return None
            else:
                for ws_name in self.sheet_names:
                    self.remove_section(section_name, ws_name)
                    section_tag_cell_existing_row_num = value_col_A.index(section_name) + 1

                    for i in range(section_tag_cell_existing_row_num, section_tag_cell_existing_row_num+6): ws.insert_rows(i)
                    section_tag_cell_existing_position = "A"+str(section_tag_cell_existing_row_num)
                    ws[section_tag_cell_existing_position] = section_name
                    header0_len = len(header0)
                    header1_len = len(header1)
                    header0_row = section_tag_cell_existing_row_num+1
                    header1_row = section_tag_cell_existing_row_num+3
                    for cell_col in range(1, header0_len+1):
                        ws.cell(row=header0_row, column=cell_col, value=header0[cell_col-1])
                    for cell_col in range(1, header1_len+1):
                        ws.cell(row=header1_row, column=cell_col, value=header1[cell_col-1])
                    End_cell_position = "A" + str(header1_row+1)
                    ws[End_cell_position] = "End"
                    self.fm.section_format(ws[section_tag_cell_existing_position], ws[End_cell_position])
                    self.fm.section_title_format(ws[str(header0_row)], ws[str(header1_row)])
        else:
            for ws_name in self.sheet_names:
                self.new_section_cell_header_create(ws_name, section_name, section_tag_cell_num, header0, header1)

    def new_section_cell_header_create(self, ws_name, section_name, section_tag_cell_num, header0, header1):
        '''
        :param ws_name:
        :param section_name:
        :param section_tag_cell_num:
        :param header0:
        :param header1:
        :return:

        Usage:
        for section tag & section title creation.
        '''
        ws = self.wb[ws_name]
        ws[section_tag_cell_num] = section_name
        section_tag_cell = ws[section_tag_cell_num]
        ws.append(header0)
        max_row1 = ws[str(ws.max_row)]
        print(max_row1)
        ws.append([])
        ws.append(header1)
        max_row2 = ws[str(ws.max_row)]
        print(max_row2)
        last_cell_num = "A" + str(ws.max_row + 1)
        ws[last_cell_num] = "End"
        section_end_tag = ws[last_cell_num]
        self.fm.section_format(section_tag_cell, section_end_tag)
        self.fm.section_title_format(max_row1, max_row2)
        self.fm.oversheet_format(ws)


    def remove_section(self, section_name, sheet_name):
        """
        remove the specified section from the master sheet.

        :param section_name: the section name
        :return: return a Non-zero value if this task was finished successfully, otherwise, return None.
        """
        ws = self.wb[sheet_name]  # load the first sheet in workbook
        colA = ws["A"]
        value_col_A = []
        for cell in colA: value_col_A.append(cell.value)
        try:
            tag_row = value_col_A.index(section_name) + 1
            end_row = 0
            for i in range(tag_row, sys.maxsize):
                if value_col_A[i] == "End":
                    end_row = i
                    break
            ws.delete_rows(tag_row, end_row-tag_row+3)
        except ValueError as e:
            print("This section("+ section_name+") cannot be found in the sheet")

    def erase_table(self, section_name, sheet_name):
        """
        erase all the data records of the specified table of the particular section.
            Note: just erase the data of the table but not the table header

        :param section_name: the section name
        :param sheet_name: the worksheet name
        :return: return the number of the rows that were erased. if it's failed, return a negative integer.
        """
        ws = self.wb[sheet_name]
        colA = ws["A"]
        value_col_A = []
        for cell in colA: value_col_A.append(cell.value)

        try:
            tag_row = value_col_A.index(section_name) + 1
            End_row = 0    #The End tag row num in section
            None_row = 0    #The None data row num in section
            for i in range(tag_row, sys.maxsize):
                if value_col_A[i] == None:
                    None_row = i+1
                    row_num = None_row - (tag_row+2)
                    #print(row_num)
                    ws.delete_rows(tag_row+2, row_num)
                    colA = ws["A"]
                    value_col_A = []
                    for cell in colA: value_col_A.append(cell.value)
                    break
            for i in range(tag_row, sys.maxsize):
                if value_col_A[i] == "End":
                    End_row = i+1
                    row_num = End_row - (tag_row+4)
                    ws.delete_rows(tag_row+4, row_num)
                    break
        except ValueError as e:
            print("This section("+ section_name+") cannot be found in the sheet")
        except IndexError as e:
            print("None Data is section")

    @property
    def section_names(self):
        """
        iterate through the master sheet and collect the section names in this master sheet.

        :return: a list that contains each section names. if there is no section, return a empty list.
        """
        ws = self.wb[self.sheet_name]
        colA = ws["A"]
        colA_list = []
        for value in colA.value: colA_list.append(value)
        return colA_list


    def get_section(self, section_name):
        """
        get the data of the specified section and return the data as a list

        :param section_name: section name
        :return: a nested list which representing the tables.  the following is an example:
            [
                # the first table in the section
                [
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                        ... ...
                    {"column1": value1, "column2": value2, ... "columnN": valueN}
                ],
                # the second table in the section
                [
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                    {"column1": value1, "column2": value2, ... "columnN": valueN},
                        ... ...
                    {"column1": value1, "column2": value2, ... "columnN": valueN}
                ]
            ]
        """
        pass

    @property
    def master_sheet(self):
        """
        get the whole data of the master sheet.
        Note: (for gathering this data, we can iterate the section_names by
            method <get_section>

        :return: a dictionary contains each section data. the following is an example:
            {
                "section1_name": section1_data,
                "section2_name": section2_data,
                    ... ...
                "sectionN_name": sectionN_data

            }

            Note: for the structure of section_data, please refer to the return value of method <get_section>
        """
        pass

    def get_table(self, section_name, table_index, sheet_name):
        """
        get the data of the specified table.
        Note: (for gathering this data, we can extract data from the return value of
            method <get_section>)

        :param section_name: the section name which the table is located in.
        :param table_index: the table index we want to get.
        :param sheet_name: Specific sheet for data
        :return: a list of dictionaries which presenting the data records, the following is an example:
             [
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                    ... ...
                {"column1": value1, "column2": value2, ... "columnN": valueN}
            ]
        """
        pass

    def insert(self, section_name, table_index, data, sheet_name):
        """
        insert data into a particular table of the specified section.

        :param section_name: section name.
        :param sheet_name: specific sheet for data insert
        :param table_index: table index. (each section only contain 2 tables with index 0, 1)
        :param data: the data need to be inserted. it should be a dictionary that presents the data record (a row)
            the structure of the data should like the following example:
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                    ... ...
                {"column1": value1, "column2": value2, ... "columnN": valueN}
        :return: the number of the rows that were inserted. if failed, raise a exception or return a negative integer.
        """
        ws = self.wb[sheet_name]
        if table_index == 0:
            pass
        elif table_index == 1:
            pass
        else:
            print("Warning: No such table!! table_index can be 0 or 1")


    def bulk_insert(self, section_name, table_index, data_list, mode="merge"):
        """
        insert bulk data into the specified table.

        :param section_name: the section name.
        :param table_index: the table index. (each section only contain 2 tables with index 0, 1)
        :param data_list: a list of dictionaries. the following is an example:
            [
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                {"column1": value1, "column2": value2, ... "columnN": valueN},
                    ... ...
                {"column1": value1, "column2": value2, ... "columnN": valueN}
            ]
        :param mode: this parameter could be "merge" or "override".
            for merge, add the data into the existing table.
            for override, erase the existing table first, then insert the data. .
        :return: the number of the rows that were inserted. if failed, raise a exception or return a negative integer.
        """
        pass
if __name__ == "__main__":
    MS = MasterSheet("C:\python-new-code\openpyxltest\stustu\\", "testplatform.xlsx")
    MS.add_section("testing_section5", ["ISE2", "Firepower2", "WLC3"], ["ise3", "Firepower", "WLC3"])
    #MS.remove_section("testing_section3", "Sheet1")
    #MS.erase_table("testing_section3", "Sheet1")
    #print(MS.section_names)
    MS.wb.save(MS.file_path + MS.sheet_name)
