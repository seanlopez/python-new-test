from openpyxl import Workbook
from openpyxl import load_workbook
import master_sheet_format


class MasterSheet(object):
    def __init__(self, file_path, sheet_name, override=False):
        self.sheet_name = sheet_name
        self.file_path = file_path
        if override == False:
            self.wb = load_workbook(file_path+sheet_name)  #Load the existing workbook
        else:
            self.wb = Workbook()    #create a new workbook in RAM

        self.fm = master_sheet_format.master_sheet_format()

    def add_section(self, section_name, header0, header1, override=True):
        sheet_names = self.wb.sheetnames
        num1_sheet = sheet_names[0]
        ws = self.wb[num1_sheet]
        colA = ws["A"]
        worksheet_start_row_num = str(len(colA)+2)
        section_tag_cell_num = "A" + worksheet_start_row_num
        value_col_A = []
        for cell in colA: value_col_A.append(cell.value)
        if section_name in value_col_A:
            if override == False:
                print("Warning! It's an existing section in worksheet")
                return None
            else:
                pass
        else:
            for ws_name in sheet_names:
                ws = self.wb[ws_name]
                ws[section_tag_cell_num] = section_name
                section_tag_cell = ws[section_tag_cell_num]
                ws.append(header0)
                max_row1 = ws[str(ws.max_row)]
                print(max_row1)
                ws.append([])
                ws.append(header1)
                max_row2 = ws[str(ws.max_row)]
                print(max_row2)
                last_cell_num = "A" + str(ws.max_row+1)
                ws[last_cell_num] = "End"
                section_end_tag = ws[last_cell_num]
                self.fm.section_format(section_tag_cell, section_end_tag)
                self.fm.section_title_format(max_row1, max_row2)
                self.fm.oversheet_format(ws)





if __name__ == "__main__":
    MS = MasterSheet("C:\python-new-code\openpyxltest\stustu\\", "testplatform.xlsx")
    MS.add_section("testing_section", ["ise", "Firepower", "WLC"], ["ise1", "Firepower2", "WLC3"])
    MS.wb.save(MS.file_path + MS.sheet_name)
