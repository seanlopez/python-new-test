from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

class master_sheet_format(object):
    def __init__(self):
        self.color_fill_tag = PatternFill(fill_type='solid', fgColor="000080")
        self.ft_tag = Font(color="FFFFFF")
        self.color_fill_title = PatternFill(fill_type='solid', fgColor="33ff33")

    def section_format(self, start_tag, end_tag):
        start_tag.font, end_tag.font = self.ft_tag, self.ft_tag
        start_tag.fill, end_tag.fill = self.color_fill_tag, self.color_fill_tag

    def section_title_format(self, header0_row, header1_row):
        for x in header0_row: x.fill = self.color_fill_title
        for x in header1_row: x.fill = self.color_fill_title

    def oversheet_format(self, worksheet):   #all column width editor
        for i in range(1, worksheet.max_column + 1): worksheet.column_dimensions[get_column_letter(i)].width = 20
