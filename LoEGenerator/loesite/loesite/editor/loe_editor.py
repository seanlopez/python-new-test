from openpyxl import load_workbook


class loe_editor(object):
    def __init__(self, form_dict, workbook_file, sheet_name):
        self.portal_form = form_dict
        self.wb = load_workbook(workbook_file)
        self.ws = self.wb[sheet_name]
        self.ws["C3"] = form_dict["customer_name"]   # fill the customer name in sheet

    def ise_requirement_phase_editor(self):
        '''
        according to the portal form value to fill in the workdays in LOE spreadsheet
        :return:
        '''
        workshop_days = 0
        document_creation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            workshop_days = workshop_days + 1
        else:
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            workshop_days = workshop_days + 1
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            workshop_days = workshop_days + 2
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            workshop_days = workshop_days + 2
            document_creation_days = document_creation_days + 1
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            workshop_days = workshop_days + 3
            document_creation_days = document_creation_days + 2
        else:
            workshop_days = workshop_days + 3
            document_creation_days = document_creation_days + 2

        if "3rdparty" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "trustsec" in self.portal_form.keys():
            workshop_days = workshop_days + 1
        if "eapfast" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5
        if "advancedguest" in self.portal_form.keys():
            workshop_days = workshop_days + 0.5

        # fill the workshop days
        self.ws["D17"] = workshop_days
        self.ws["E17"] = workshop_days

        # fill the document creation days
        self.ws["D18"] = document_creation_days
        self.ws["E18"] = document_creation_days

    def ise_design_phase_editor(self):
        '''
        according to the portal form value to fill in the workdays in LOE spreadsheet
        :return:
        '''
        workshop_days = 1
        design_document_days = 0
        if self.portal_form["sdaornot"] == "yes":
            workshop_days = workshop_days + 0

        if self.portal_form["deploymentmethod"] == "aaa":
            design_document_days = design_document_days + 4
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            design_document_days = design_document_days + 4.5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            design_document_days = design_document_days + 5
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            workshop_days = workshop_days + 1
            design_document_days = design_document_days + 5.5
        else:
            workshop_days = workshop_days + 1
            design_document_days = design_document_days + 6

        if "3rdparty" in self.portal_form.keys():
            design_document_days = design_document_days + 0.5
        if "trustsec" in self.portal_form.keys():
            design_document_days = design_document_days + 1
        if "eapfast" in self.portal_form.keys():
            design_document_days = design_document_days + 1
        if "advancedguest" in self.portal_form.keys():
            design_document_days = design_document_days + 0.5

        # fill the workshop days
        self.ws["D21"] = workshop_days
        self.ws["E21"] = workshop_days

        # fill the document creation days
        self.ws["D22"] = design_document_days
        self.ws["E22"] = design_document_days

    def ise_nip_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        documentation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            documentation_days = documentation_days + 0

        if self.portal_form["deploymentmethod"] == "aaa":
            documentation_days = documentation_days + 4
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            documentation_days = documentation_days + 5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            documentation_days = documentation_days + 6
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            documentation_days = documentation_days + 7
        else:
            documentation_days = documentation_days + 8

        if "3rdparty" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5
        if "trustsec" in self.portal_form.keys():
            documentation_days = documentation_days + 1
        if "eapfast" in self.portal_form.keys():
            documentation_days = documentation_days + 1
        if "advancedguest" in self.portal_form.keys():
            documentation_days = documentation_days + 0.5

        # fill the document creation days
        self.ws["D29"] = documentation_days
        self.ws["E29"] = documentation_days

    def ise_nruf_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        documentation_days = 0
        if self.portal_form["sdaornot"] == "yes":
            documentation_days = documentation_days + 0

        if self.portal_form["deploymentmethod"] == "aaa":
            documentation_days = documentation_days + 4
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            documentation_days = documentation_days + 5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            documentation_days = documentation_days + 5
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            documentation_days = documentation_days + 5
        else:
            documentation_days = documentation_days + 5

        if "3rdparty" in self.portal_form.keys():
            documentation_days = documentation_days + 0
        if "trustsec" in self.portal_form.keys():
            documentation_days = documentation_days + 0
        if "eapfast" in self.portal_form.keys():
            documentation_days = documentation_days + 0
        if "advancedguest" in self.portal_form.keys():
            documentation_days = documentation_days + 0

        # fill the document creation days
        self.ws["D35"] = documentation_days
        self.ws["E35"] = documentation_days

        if "autotest" in self.portal_form.keys():
            self.ws["D38"] = 10

    def ise_lab_testing_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        lab_building_days = 0
        test_execution_days = 0

        if self.portal_form["sdaornot"] == "yes":
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            lab_building_days = lab_building_days + 0
            test_execution_days = test_execution_days + 1
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1.5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            lab_building_days = lab_building_days + 1
            test_execution_days = test_execution_days + 1.5
            self.ws["D44"] = 1
            self.ws["E44"] = 1
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            lab_building_days = lab_building_days + 2
            test_execution_days = test_execution_days + 2
            test_execution_days = test_execution_days + 2.5
            self.ws["D44"] = 1
            self.ws["E44"] = 1
        else:
            lab_building_days = lab_building_days + 2.5
            test_execution_days = test_execution_days + 2.5
            self.ws["D44"] = 1
            self.ws["E44"] = 1

        if "3rdparty" in self.portal_form.keys():
            self.ws["D45"] = 0.5
            self.ws["E45"] = 0.5
        if "datamig" in self.portal_form.keys():
            self.ws["D46"] = 0.5
            self.ws["E46"] = 0.5
        if "auto" in self.portal_form.keys():
            self.ws["D47"] = 2
            self.ws["E47"] = 2

        # lab building days
        self.ws["D42"] = lab_building_days
        self.ws["E42"] = lab_building_days

        # testing days
        self.ws["D43"] = test_execution_days
        self.ws["E43"] = test_execution_days

    def ise_implementation_phase_editor(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        if self.portal_form["sdaornot"] == "yes":
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            self.ws["D52"] = 3
            self.ws["E52"] = 3
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            self.ws["D52"] = 3
            self.ws["E52"] = 3
            self.ws["D53"] = 0.5
            self.ws["E53"] = 0.5
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            self.ws["D52"] = 3
            self.ws["E52"] = 3
            self.ws["D54"] = 0.5
            self.ws["E54"] = 0.5
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            self.ws["D52"] = 3
            self.ws["E52"] = 3
            self.ws["D54"] = 0.5
            self.ws["E54"] = 0.5
            self.ws["D55"] = 0.5
            self.ws["E55"] = 0.5
        else:
            self.ws["D52"] = 3
            self.ws["E52"] = 3
            self.ws["D53"] = 0.5
            self.ws["E53"] = 0.5
            self.ws["D54"] = 0.5
            self.ws["E54"] = 0.5
            self.ws["D55"] = 0.5
            self.ws["E55"] = 0.5

        if "3rdparty" in self.portal_form.keys():
            self.ws["D57"] = 0.5
            self.ws["E57"] = 0.5
        if "advanceguest" in self.portal_form.keys():
            self.ws["D46"] = 0.5
            self.ws["E46"] = 0.5

    def ise_kt_phase(self):
        '''
            according to the portal form value to fill in the workdays in LOE spreadsheet
            :return:
        '''
        kt_document_days = 0
        training_days = 0
        if self.portal_form["sdaornot"] == "yes":
            pass

        if self.portal_form["deploymentmethod"] == "aaa":
            kt_document_days = 3
            training_days = 0.5
        elif self.portal_form["deploymentmethod"] == "advancedNAC":
            kt_document_days = 4
            training_days = 1
        elif self.portal_form["deploymentmethod"] == "simplebyod":
            kt_document_days = 4
            training_days = 1
        elif self.portal_form["deploymentmethod"] == "byodNAC":
            kt_document_days = 4
            training_days = 1
        else:
            kt_document_days = 5
            training_days = 2

        # fill in the kt document days
        self.ws["D62"] = kt_document_days
        self.ws["E62"] = kt_document_days

        # fill in the training days
        self.ws["D63"] = training_days
        self.ws["E63"] = training_days

    def save_close_sheet(self, output_path):
        output_file_name = f"{self.portal_form['customer_name']}_Security_LoE.xlsx"
        self.wb.save(output_path + "/" + f"{self.portal_form['customer_name']}_Security_LoE.xlsx")
        return output_file_name

#wb.save("../output_LoE/test.xlsx")
